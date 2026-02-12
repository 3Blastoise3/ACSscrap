#!/usr/bin/env python3
"""
==============================================================================
AUTOMATED CENSUS ACS DATA SCRAPER
==============================================================================

SETUP & INSTALLATION:
    pip install requests openpyxl pandas

USAGE:
    python3 scrape_acs.py
~
OUTPUT:
    Creates 'redbook_acs_output.xlsx' in the current directory with one
    sheet per dataset, formatted to match the working spreadsheet.

CONFIGURATION:
    Edit the CONFIG section below to change:
    - Year (auto-calculated as current year - 2, falls back if unavailable)
    - Survey type (acs1 or acs5)
    - API key (optional, 500 requests/day without key)
    - Output filename

TABLES PROCESSED:
    RB002 - Age Group by % of Population (S0101)
    RB032 - Educational Attainment of Pop 25+ (B15002)
    RB039 - Metropolitan Commuting (B08303)
    RB040 - % Workers Who Worked From Home (C08301/S0801)
    RB044 - % Without Health Insurance (S2702)

==============================================================================
"""

import requests
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import time
from datetime import datetime
from typing import Dict, List, Tuple, Optional

# ==============================================================================
# CONFIGURATION
# ==============================================================================

# Calculate most recent expected ACS year (typically 2 years behind current year)
# The script will automatically fall back to earlier years if data isn't available
CURRENT_YEAR = datetime.now().year
MOST_RECENT_ACS_YEAR = CURRENT_YEAR - 2

CONFIG = {
    'year': MOST_RECENT_ACS_YEAR,  # Auto-calculated, will fall back to earlier years if needed
    'survey_type': 'acs1',          # 'acs1' for 1-year or 'acs5' for 5-year
    'api_key': '',                  # Optional - leave empty for 500 requests/day
    'output_file': 'redbook_acs_output.xlsx',
    'max_retries': 3,
    'retry_delay': 2,               # seconds
}

# ==============================================================================
# ACS TABLE DEFINITIONS
# ==============================================================================

TABLES = {
    'RB002': {
        'table_id': 'S0101',
        'name': 'Age Group by % of Population',
        'sheet_name': 'Age Group by % of Population',
        'geography': 'state:*',
        'source_note': 'Source: American Community Survey, U.S. Census Bureau'
    },
    'RB032': {
        'table_id': 'S1501',
        'name': 'Educational Attainment of Pop 25+',
        'sheet_name': 'Educ Attnment of Pop 25 & older',
        'geography': 'state:*',
        'source_note': 'Source: American Community Survey'
    },
    'RB039': {
        'table_id': 'S0801',
        'name': 'Metropolitan Commuting - Mean Travel Time',
        'sheet_name': 'Metropolitan Commuting',
        'geography': 'metropolitan statistical area/micropolitan statistical area:*',
        'source_note': 'Source: U.S. Census Bureau'
    },
    'RB039B': {
        'table_id': 'S0801',
        'name': 'Metropolitan Commuting - Mode of Transportation',
        'sheet_name': 'Mode of Transportation',
        'geography': 'metropolitan statistical area/micropolitan statistical area:*',
        'source_note': 'Source: U.S. Census Bureau'
    },
    'RB040': {
        'table_id': 'S0801',
        'name': '% Workers Who Worked From Home',
        'sheet_name': '% Workers who work from home',
        'geography': 'metropolitan statistical area/micropolitan statistical area:*',
        'source_note': 'Source: U.S. Census Bureau'
    },
    'RB044': {
        'table_id': 'S2701',
        'name': '% Without Health Insurance',
        'sheet_name': '% WO Health Insurance by State',
        'geography': 'state:*',
        'source_note': 'Source: American Community Survey'
    }
}

# ==============================================================================
# CENSUS API HELPER FUNCTIONS
# ==============================================================================

def get_dataset_path(table_id: str, survey_type: str) -> str:
    """Determine the API dataset path based on table prefix."""
    prefix = table_id[0]

    if prefix == 'S':
        return f"{survey_type}/subject"
    elif prefix == 'D' and table_id.startswith('DP'):
        return f"{survey_type}/profile"
    elif prefix == 'C' and table_id.startswith('CP'):
        return f"{survey_type}/cprofile"
    else:
        return survey_type


def fetch_variable_labels(table_id: str, year: int, dataset_path: str) -> Dict:
    """Fetch variable labels/metadata from Census API."""
    url = f"https://api.census.gov/data/{year}/acs/{dataset_path}/groups/{table_id}.json"

    try:
        response = requests.get(url, timeout=30)
        if response.ok:
            data = response.json()
            return data.get('variables', {})
    except Exception as e:
        print(f"  Warning: Could not fetch variable labels: {e}")

    return {}


def clean_label(label: str) -> str:
    """Clean up Census variable labels."""
    # Remove common prefixes
    label = label.replace('Estimate!!', '')
    label = label.replace('!!Estimate', '')
    label = label.replace('Total!!', '')

    # Replace !! with > for hierarchy
    label = label.replace('!!', ' > ')

    return label.strip()


def fetch_acs_data(table_id: str, year: int, survey_type: str, geography: str, api_key: str = '') -> Tuple[Optional[List], Optional[Dict]]:
    """
    Fetch ACS data from Census API.

    Returns:
        Tuple of (data_rows, variable_labels) or (None, None) on error
    """
    dataset_path = get_dataset_path(table_id, survey_type)

    # Fetch variable labels first
    var_labels = fetch_variable_labels(table_id, year, dataset_path)

    # Build data URL
    url = f"https://api.census.gov/data/{year}/acs/{dataset_path}?get=group({table_id})&for={geography}"
    if api_key:
        url += f"&key={api_key}"

    # Retry logic
    for attempt in range(CONFIG['max_retries']):
        try:
            response = requests.get(url, timeout=30)

            if response.ok:
                data = response.json()
                if data and len(data) > 1:
                    return data, var_labels
            elif response.status_code == 404:
                # Try previous year
                if attempt == 0 and year > 2020:
                    print(f"  Table not found for {year}, trying {year-1}...")
                    year -= 1
                    url = url.replace(f"/{year+1}/", f"/{year}/")
                    continue

            print(f"  API Error: Status {response.status_code}")

        except requests.exceptions.RequestException as e:
            print(f"  Network error (attempt {attempt + 1}/{CONFIG['max_retries']}): {e}")
            if attempt < CONFIG['max_retries'] - 1:
                time.sleep(CONFIG['retry_delay'])

    return None, None


def find_variable_by_label(var_labels: Dict, search_terms: List[str], exclude_terms: List[str] = None) -> Optional[str]:
    """Find a variable code by searching for terms in its label, optionally excluding terms."""
    if exclude_terms is None:
        exclude_terms = []

    for var_code, var_info in var_labels.items():
        if not var_code.endswith('E'):  # Only estimate columns
            continue

        label = var_info.get('label', '').lower()

        # Check if all search terms are in the label
        if not all(term.lower() in label for term in search_terms):
            continue

        # Check if any exclude terms are in the label
        if any(term.lower() in label for term in exclude_terms):
            continue

        return var_code

    return None


# ==============================================================================
# DATA PROCESSING FUNCTIONS
# ==============================================================================

def process_rb002_age_groups(data: List, var_labels: Dict) -> List[Dict]:
    """
    Process S0101 - Age Group by % of Population

    Output columns: State | Rank | %<18 | %18-24 | %25-44 | %45-64 | %>64
    """
    headers = data[0]
    rows = data[1:]

    # Find variable indices for age groups (looking for percentage of TOTAL population, not male/female)
    # Exclude "male" and "female" to get the C02 column (Percent!!) instead of C04/C06 (Percent Male/Female)
    exclude = ['male', 'female']

    # These age groups exist as single categories
    age_groups = {
        '<18': find_variable_by_label(var_labels, ['percent', 'total population', 'under 18'], exclude),
        '18-24': find_variable_by_label(var_labels, ['percent', 'total population', '18 to 24'], exclude),
        '>64': find_variable_by_label(var_labels, ['percent', 'total population', '65 years and over'], exclude)
    }

    # These need to be calculated by summing 5-year age ranges
    age_25_29 = find_variable_by_label(var_labels, ['percent', 'total population', '25 to 29'], exclude)
    age_30_34 = find_variable_by_label(var_labels, ['percent', 'total population', '30 to 34'], exclude)
    age_35_39 = find_variable_by_label(var_labels, ['percent', 'total population', '35 to 39'], exclude)
    age_40_44 = find_variable_by_label(var_labels, ['percent', 'total population', '40 to 44'], exclude)
    age_45_49 = find_variable_by_label(var_labels, ['percent', 'total population', '45 to 49'], exclude)
    age_50_54 = find_variable_by_label(var_labels, ['percent', 'total population', '50 to 54'], exclude)
    age_55_59 = find_variable_by_label(var_labels, ['percent', 'total population', '55 to 59'], exclude)
    age_60_64 = find_variable_by_label(var_labels, ['percent', 'total population', '60 to 64'], exclude)

    # Get column indices
    name_idx = headers.index('NAME') if 'NAME' in headers else 0

    result = []
    for row in rows:
        state_name = row[name_idx]

        row_data = {'State': state_name}

        # Process simple age groups
        for age_label, var_code in age_groups.items():
            if var_code and var_code in headers:
                idx = headers.index(var_code)
                value = row[idx]
                try:
                    row_data[f'%{age_label}'] = float(value) if value else None
                except:
                    row_data[f'%{age_label}'] = None

        # Calculate 25-44 by summing individual age ranges
        sum_25_44 = 0
        for var_code in [age_25_29, age_30_34, age_35_39, age_40_44]:
            if var_code and var_code in headers:
                idx = headers.index(var_code)
                try:
                    sum_25_44 += float(row[idx]) if row[idx] else 0
                except:
                    pass
        row_data['%25-44'] = sum_25_44 if sum_25_44 > 0 else None

        # Calculate 45-64 by summing individual age ranges
        sum_45_64 = 0
        for var_code in [age_45_49, age_50_54, age_55_59, age_60_64]:
            if var_code and var_code in headers:
                idx = headers.index(var_code)
                try:
                    sum_45_64 += float(row[idx]) if row[idx] else 0
                except:
                    pass
        row_data['%45-64'] = sum_45_64 if sum_45_64 > 0 else None

        result.append(row_data)

    # Sort by %<18 descending and add rank
    result.sort(key=lambda x: x.get('%<18', 0) or 0, reverse=True)
    for i, row in enumerate(result, 1):
        row['Rank'] = i

    return result


def process_rb032_education(data: List, var_labels: Dict) -> List[Dict]:
    """
    Process S1501 - Educational Attainment of Pop 25+

    Output: Rank | State | Completed H.S. or Higher | Bachelors or Higher | Rank | Advanced Degree | Rank
    """
    headers = data[0]
    rows = data[1:]

    name_idx = headers.index('NAME') if 'NAME' in headers else 0

    # Find percent columns for education levels (population 25+)
    # Exclude male/female to get C02 column (total population, not by gender)
    # Exclude age-specific groups like "25 to 34" to get all "25 years and over"
    exclude = ['male', 'female', '25 to 34', '35 to 44', '45 to 64', '65 years and over', 'earnings', 'median']
    hs_var = find_variable_by_label(var_labels, ['high school graduate or higher', 'percent', '25 years and over'], exclude)
    bachelors_var = find_variable_by_label(var_labels, ['bachelor', 'or higher', 'percent', '25 years and over'], exclude)
    advanced_var = find_variable_by_label(var_labels, ['graduate', 'professional', 'percent', '25 years and over'], exclude)

    result = []
    for row in rows:
        state_name = row[name_idx]

        row_data = {'State': state_name}

        # Extract percentages
        if hs_var and hs_var in headers:
            idx = headers.index(hs_var)
            try:
                row_data['Completed H.S. or Higher'] = float(row[idx]) if row[idx] else None
            except:
                row_data['Completed H.S. or Higher'] = None
        else:
            row_data['Completed H.S. or Higher'] = None

        if bachelors_var and bachelors_var in headers:
            idx = headers.index(bachelors_var)
            try:
                row_data['Bachelors or Higher'] = float(row[idx]) if row[idx] else None
            except:
                row_data['Bachelors or Higher'] = None
        else:
            row_data['Bachelors or Higher'] = None

        if advanced_var and advanced_var in headers:
            idx = headers.index(advanced_var)
            try:
                row_data['Advanced Degree'] = float(row[idx]) if row[idx] else None
            except:
                row_data['Advanced Degree'] = None
        else:
            row_data['Advanced Degree'] = None

        result.append(row_data)

    # Sort by "Completed H.S. or Higher" descending and add main rank
    result.sort(key=lambda x: x.get('Completed H.S. or Higher', 0) or 0, reverse=True)
    for i, row in enumerate(result, 1):
        row['Rank'] = i

    # Add separate ranks for Bachelors and Advanced Degree
    sorted_by_bachelors = sorted(result, key=lambda x: x.get('Bachelors or Higher', 0) or 0, reverse=True)
    for i, row in enumerate(sorted_by_bachelors, 1):
        row['Bachelors Rank'] = i

    sorted_by_advanced = sorted(result, key=lambda x: x.get('Advanced Degree', 0) or 0, reverse=True)
    for i, row in enumerate(sorted_by_advanced, 1):
        row['Advanced Rank'] = i

    return result


def process_rb039_commuting(data: List, var_labels: Dict) -> List[Dict]:
    """
    Process S0801 - Metropolitan Commuting - Mean Travel Time

    Output: Rank | Metro Area | Average Commute Time
    """
    headers = data[0]
    rows = data[1:]

    name_idx = headers.index('NAME') if 'NAME' in headers else 0

    # Find mean travel time variable from total population (exclude male/female)
    exclude = ['male', 'female']
    mean_var = find_variable_by_label(var_labels, ['mean travel time', 'workers 16 years'], exclude)

    result = []
    for row in rows:
        metro_name = row[name_idx]

        avg_commute = None
        if mean_var and mean_var in headers:
            idx = headers.index(mean_var)
            try:
                avg_commute = float(row[idx]) if row[idx] else None
            except:
                pass

        result.append({
            'Metro Area': metro_name,
            'Average Commute Time': avg_commute
        })

    # Sort by commute time descending and add rank
    result.sort(key=lambda x: x.get('Average Commute Time', 0) or 0, reverse=True)
    for i, row in enumerate(result, 1):
        row['Rank'] = i

    return result


def process_rb039b_mode_of_transportation(data: List, var_labels: Dict) -> List[Dict]:
    """
    Process S0801 - Mode of Transportation

    Output: Rank | Metro Area | % Drove Alone | % Carpooled | % Public Transit
    """
    headers = data[0]
    rows = data[1:]

    name_idx = headers.index('NAME') if 'NAME' in headers else 0

    # Find mode of transportation variables from total population (exclude male/female)
    # For carpooled, exclude the subcategories (2-person, 3-person, 4-or-more) to get the total
    exclude = ['male', 'female', '2-person', '3-person', '4-or-more']
    drove_alone_var = find_variable_by_label(var_labels, ['car, truck, or van', 'drove alone', 'workers 16 years'], exclude)
    carpooled_var = find_variable_by_label(var_labels, ['car, truck, or van', 'carpooled', 'workers 16 years'], exclude)
    public_transit_var = find_variable_by_label(var_labels, ['public transportation', 'workers 16 years'], exclude)

    result = []
    for row in rows:
        metro_name = row[name_idx]

        row_data = {'Metro Area': metro_name}

        # Extract mode of transportation percentages
        if drove_alone_var and drove_alone_var in headers:
            idx = headers.index(drove_alone_var)
            try:
                row_data['% Drove Alone'] = float(row[idx]) if row[idx] else None
            except:
                row_data['% Drove Alone'] = None

        if carpooled_var and carpooled_var in headers:
            idx = headers.index(carpooled_var)
            try:
                row_data['% Carpooled'] = float(row[idx]) if row[idx] else None
            except:
                row_data['% Carpooled'] = None

        if public_transit_var and public_transit_var in headers:
            idx = headers.index(public_transit_var)
            try:
                row_data['% Public Transit'] = float(row[idx]) if row[idx] else None
            except:
                row_data['% Public Transit'] = None

        result.append(row_data)

    # Sort by % Public Transit descending and add rank
    result.sort(key=lambda x: x.get('% Public Transit', 0) or 0, reverse=True)
    for i, row in enumerate(result, 1):
        row['Rank'] = i

    return result


def process_rb040_wfh(data: List, var_labels: Dict) -> List[Dict]:
    """
    Process S0801 - % Workers Who Worked From Home

    Output: Rank | Metro Area | 2024 | Rank | 2021 | Rank | 2019 | Rank
    """
    headers = data[0]
    rows = data[1:]

    name_idx = headers.index('NAME') if 'NAME' in headers else 0

    # Find work from home percentage variable from total population (exclude male/female)
    exclude = ['male', 'female']
    wfh_var = find_variable_by_label(var_labels, ['worked from home', 'workers 16 years'], exclude)

    result = []
    for row in rows:
        metro_name = row[name_idx]

        wfh_pct = None
        if wfh_var and wfh_var in headers:
            idx = headers.index(wfh_var)
            try:
                wfh_pct = float(row[idx]) if row[idx] else None
            except:
                pass

        result.append({
            'Metro Area': metro_name,
            str(CONFIG['year']): wfh_pct  # Current year only for now
        })

    # Sort and add rank
    result.sort(key=lambda x: x.get(str(CONFIG['year']), 0) or 0, reverse=True)
    for i, row in enumerate(result, 1):
        row['Rank'] = i

    return result


def process_rb044_health_insurance(data: List, var_labels: Dict) -> List[Dict]:
    """
    Process S2701 - % Without Health Insurance

    Output: Rank | State | Percent Uninsured | Age <19 | Rank | Age 19-64 | Rank | Age 65+ | Rank
    """
    headers = data[0]
    rows = data[1:]

    name_idx = headers.index('NAME') if 'NAME' in headers else 0

    # Find uninsured percentage variables by age group from S2701 C05 column (Percent Uninsured)
    # For total, we want S2701_C05_001E which is the overall population, not subcategories
    uninsured_vars = {
        'Total': 'S2701_C05_001E',  # Direct code for overall total percent uninsured
        '<19': find_variable_by_label(var_labels, ['percent uninsured', 'under 19 years']),
        '19-64': find_variable_by_label(var_labels, ['percent uninsured', '19 to 64 years']),
        '65+': find_variable_by_label(var_labels, ['percent uninsured', '65 years and older'])
    }

    result = []
    for row in rows:
        state_name = row[name_idx]

        row_data = {'State': state_name}

        for age_label, var_code in uninsured_vars.items():
            col_name = f'Age {age_label}' if age_label != 'Total' else 'Percent Uninsured'

            if var_code and var_code in headers:
                idx = headers.index(var_code)
                try:
                    row_data[col_name] = float(row[idx]) if row[idx] else None
                except:
                    row_data[col_name] = None

        result.append(row_data)

    # Sort by total percent uninsured ascending (lowest first) and add ranks
    result.sort(key=lambda x: x.get('Percent Uninsured', 100) or 100)
    for i, row in enumerate(result, 1):
        row['Rank'] = i

    # Add age group ranks
    for age_col in ['Age <19', 'Age 19-64', 'Age 65+']:
        if age_col in result[0]:
            sorted_by_age = sorted(result, key=lambda x: x.get(age_col, 100) or 100)
            for i, row in enumerate(sorted_by_age, 1):
                row[f'{age_col} Rank'] = i

    return result


# ==============================================================================
# EXCEL OUTPUT FUNCTIONS
# ==============================================================================

def write_to_excel(all_data: Dict[str, List[Dict]], output_file: str):
    """Write processed data to Excel with formatting."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for table_id, table_info in TABLES.items():
        if table_id not in all_data or not all_data[table_id]:
            continue

        sheet_name = table_info['sheet_name']
        data = all_data[table_id]

        ws = wb.create_sheet(title=sheet_name)

        # Write headers
        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Identify percentage columns (columns that contain percentage data)
        percentage_columns = set()
        for header in headers:
            # Check if header indicates a percentage
            if (header and (
                header.startswith('%') or
                'Percent' in header or
                header in ['Completed H.S. or Higher', 'Bachelors or Higher', 'Advanced Degree'] or
                header.isdigit()  # Year columns in WFH sheet (2024, 2021, etc.)
            ) and header not in ['Rank', 'State', 'Metro Area']):  # Exclude non-percentage columns
                percentage_columns.add(header)

        # Write data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Format percentage columns with % symbol (values are already 0-100 scale)
                if header in percentage_columns and isinstance(value, (int, float)):
                    cell.number_format = '0.0"%"'  # Custom format: shows number with % but doesn't multiply by 100

        # Add source note in far right column
        source_col = len(headers) + 2
        ws.cell(row=4, column=source_col, value=table_info['source_note'])
        ws.cell(row=4, column=source_col).font = Font(italic=True, size=9)

        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

    wb.save(output_file)


# ==============================================================================
# MAIN EXECUTION
# ==============================================================================

def main():
    """Main execution function."""
    print("=" * 80)
    print("AUTOMATED CENSUS ACS DATA SCRAPER")
    print("=" * 80)
    print(f"\nConfiguration:")
    print(f"  Year: {CONFIG['year']}")
    print(f"  Survey: {CONFIG['survey_type'].upper()}")
    print(f"  Output: {CONFIG['output_file']}")
    print()

    all_data = {}
    success_count = 0
    fail_count = 0
    failed_tables = []

    total_tables = len(TABLES)

    for idx, (table_id, table_info) in enumerate(TABLES.items(), 1):
        print(f"[{idx}/{total_tables}] Fetching {table_id} – {table_info['name']} ({table_info['table_id']})...", end=' ')

        # Fetch data from API
        data, var_labels = fetch_acs_data(
            table_info['table_id'],
            CONFIG['year'],
            CONFIG['survey_type'],
            table_info['geography'],
            CONFIG['api_key']
        )

        if data is None:
            print("[FAIL] FAILED")
            fail_count += 1
            failed_tables.append(f"{table_id} - {table_info['name']}")
            continue

        # Process data based on table type
        try:
            if table_id == 'RB002':
                processed = process_rb002_age_groups(data, var_labels)
            elif table_id == 'RB032':
                processed = process_rb032_education(data, var_labels)
            elif table_id == 'RB039':
                processed = process_rb039_commuting(data, var_labels)
            elif table_id == 'RB039B':
                processed = process_rb039b_mode_of_transportation(data, var_labels)
            elif table_id == 'RB040':
                processed = process_rb040_wfh(data, var_labels)
            elif table_id == 'RB044':
                processed = process_rb044_health_insurance(data, var_labels)
            else:
                processed = []

            all_data[table_id] = processed
            print(f"[OK] done ({len(processed)} rows)")
            success_count += 1

        except Exception as e:
            print(f"[ERROR] Processing error: {e}")
            fail_count += 1
            failed_tables.append(f"{table_id} - {table_info['name']}")

    # Write to Excel
    if all_data:
        print(f"\nWriting output to {CONFIG['output_file']}...", end=' ')
        try:
            write_to_excel(all_data, CONFIG['output_file'])
            print("[OK] done")
        except Exception as e:
            print(f"[FAIL] Error writing Excel: {e}")
            return

    # Summary
    print("\n" + "=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print(f"Success: {success_count}/{total_tables}")
    print(f"Failed:  {fail_count}/{total_tables}")

    if failed_tables:
        print("\nFailed tables:")
        for table in failed_tables:
            print(f"  • {table}")

    if all_data:
        print(f"\n[OK] Output saved to: {CONFIG['output_file']}")
    else:
        print("\n[FAIL] No data was successfully processed")

    print("=" * 80)


if __name__ == '__main__':
    main()
